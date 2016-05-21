"""
Skew-T Log-P Diagram Program

This program defines a function skewt which graphs a Skew-T Log P diagram given
a list of pressure levels, temperatures, and dew points. Before defining that
function, a method is given to extract pressure levels, temperatures, and dew
points from an Excel file that contains those values. Note that some specific
values of the mixing ratio lines are found the the Excel file,
'Math_for_Skew_T.xlsx'.

Note: Here is a link to the NOAA radiosonde realtime database (for data):
http://www.esrl.noaa.gov/raobs/

Shawn Murdzek
sfm5282@psu.edu
"""


from matplotlib import pyplot as plt
import openpyxl as op
from pylab import axes
import numpy as np
import math
import thermo_scripts as ts


# Extracting sample data from Excel using openpyxl
# Note that for columns, 1 means A, 2 means B, etc.
wb = op.load_workbook('koun_sounding.xlsx')                    
sheet = wb.get_sheet_by_name('Sheet')      
prs = np.zeros([62], 'd')
temp = np.zeros([62], 'd')
dew = np.zeros([62], 'd')
for i in xrange(5, 67):                                         
    prs[i - 5] = float(sheet.cell(row = i, column = 1).value)
    temp[i - 5] = float(sheet.cell(row = i, column = 3).value)
    dew[i - 5] = float(sheet.cell(row = i, column = 4).value)


# Function to "skew the T"
def skewing_t(T, P):
    """
    Takes the logarithm of the pressure and applies a transformation to the
    temperature so that it appears to be skewed 45 degrees. The transformation
    is T* = T + 80(3-log(P)), where T* is the temperature plotted by Python and
    T is the actual temperature.
    Inputs:
        T = 1D array of temperatures (deg C)
        P = 1D array of pressures (mb)
    Outputs:
        skwt_array = 1D array of temperatures transformed to appear skewed
        logP = 1D array of the log(P)
    """
    
    skwt_array = np.zeros([len(T)], 'd')
    logP = np.log10(P)
    for i in xrange(len(T)):
        skwt_array[i] = T[i] + 80 * (3 - logP[i])
        
    return skwt_array, logP


# Graphing the data:
def skewt(prs, temp, dew, mixing = 'on', dry_adiabat = 'on',
          moist_adiabat = 'on', show_points = 'off', parcel = 'sfc'):
    """
    This function plots temperatures and dew points from a sounding on a Skew-T
    Log-P diagram.
    Inputs:
        prs = Pressure 1D array (mb)
        temp = Temperature 1D array (deg C)
        dew = Dewpoint temperatures 1D array (deg C)
    Keywords:
        mixing = Controls whether mixing ratio lines will be included in the
            Skew-T.
        dry_adiabats = Controls whether the dry adiabats will be included in the
            Skew_T.
        moist_adiabats = Controls whether moist adiabats will be included in the
            Skew-T.
        show_points = Controls whether each individual point will be marked with
            a marker.
        parcel = Controls where the parcel drawn on the Skew-T will originate.
            'sfc' means the parcel originates at the surface, numerical values
            can also be chosen between the minimum and maximum pressure levels.
            Setting parcel = 'off' means no parcel will be drawn.
    """

    # Convert lists to arrays:
    prs = np.asarray(prs)
    temp = np.asarray(temp)
    dew = np.asarray(dew)

    # Plot the temperatures and dew points:
    # Note that each ax1 is a different set of lines
    [skwt, logP] = skewing_t(temp, prs) 
    [skwd, logP] = skewing_t(dew, prs)
    fig = plt.figure()
    ax1 = fig.add_subplot(111)
    if show_points == 'on':                                     
        ax1.plot(skwt, logP, 'r-', lw = 1.5, marker = '.',      
                 markersize = 10.0, markerfacecolor = 'r')      
        ax1.plot(skwd, logP, 'g-', lw = 1.5, marker = '.',
                 markersize = 10.0, markerfacecolor = 'g')
    else:
        ax1.plot(skwt, logP, 'r-', skwd, logP, 'g-', lw = 1.5)
        
    # Plot the isotherms:
    for i in xrange(15):                                        
        x = [30 - (10 * i), 40]
        y = [3, 3 - ((i + 1) * 0.125)]
        ax1.plot(x, y, 'k-', lw = 0.75)

    # Plot mixing ratio lines:
    if mixing == 'on':
        print "Graphing Mixing Ratio Lines..."
        mix_r = np.array([28, 18, 12, 8, 5, 3, 1.5, 0.6, 0.3], 'd')
        for i in xrange(len(mix_r)):
            mix_line = ts.mix_ratio(mix_r[i-1] * 0.001)
            [skw_mix_t, mix_p_log] = skewing_t(mix_line[:, 1] - 273.15,
                                               mix_line[:, 0] * 0.01)
            ax1.plot(skw_mix_t, mix_p_log, 'b--', lw = 0.75)

    # Plot dry adiabats                                      
    if dry_adiabat == 'on':
        print "Graphing Dry Adiabats..."
        theta = np.arange(-30, 150, 10, 'd')
        for i in xrange(len(theta)):
            dry_adiabat = ts.DALR(theta[i] + 273.15)
            [skw_ad_t, adiabat_p_log] = skewing_t(dry_adiabat[:, 1] - 273.15,
                                                  dry_adiabat[:, 0] * 0.01)
            ax1.plot(skw_ad_t, adiabat_p_log, 'k--', lw = 0.75)

    # Plot moist adiabats
    if moist_adiabat == 'on':
        print "Graphing Moist Adiabats..."
        theta_e = np.arange(-20, 150, 15)
        for i in xrange(theta_e.size):
            moist_adiabats = ts.MALR(theta_e[i] + 273.15)
            [skw_m_ad_t, m_ad_p_log] = skewing_t(moist_adiabats[:, 1] - 273.15,
                                                 moist_adiabats[:, 0]*0.01)
            ax1.plot(skw_m_ad_t, m_ad_p_log, 'm--', lw = 0.75)

    # Plot parcel
    if (parcel == 'sfc') or ((parcel >= np.amin(prs)) and (parcel <=
                                                         np.amax(prs))):
        # P, T, and D are the pressure, temperature, and dew point at the level
        # where the parcel originates
        print "Plotting Parcel..."
        if parcel == 'sfc':
            P = np.amax(prs)
        elif (parcel >= np.amin(prs)) or (parcel <= np.amax(prs)):
            P = parcel
        [p_prs, p_temp, p_dew] = ts.parcel_prof(P * 100.0, prs * 100.0,
                                                temp + 273.15, dew + 273.15,
                                                step = 1000, p1 = P * 100.0)
        [skw_p_temp, log_p_prs] = skewing_t(p_temp - 273.15, p_prs * 0.01)
        [skw_p_dew, log_p_prs] = skewing_t(p_dew - 273.15, p_prs * 0.01)
        ax1.plot(skw_p_temp, log_p_prs, 'k', lw = 1.5)
        ax1.plot(skw_p_dew, log_p_prs, 'k', lw = 1.5)
        
    # Make y-axis logarithmic, labels, title        
    plt.axis([-40, 40, 3, 2])                                    
    axis = np.log10(np.arange(100, 1100, 100))                  
    plt.yticks(axis, xrange(100, 1100, 100))    
    axes().yaxis.grid(True)                                     
    plt.xlabel('Temperature and Dew Point (deg C)')
    plt.ylabel('Pressure (mb)')
    plt.title('Skew-T Log P Diagram', size = 20.0)
    plt.show()                                                  


print ts.EL(prs * 100.0, temp + 273.15, dew + 273.15)
print ts.LFC(prs * 100.0, temp + 273.15, dew + 273.15)
skewt(prs, temp, dew, parcel = 'sfc')
#skewt([900, 850, 500], [10, 5, -20], [0, -15, -35], show_points = 'on',
#      parcel = 'sfc')


def blank_skewt():
    """
    This function simply creates a blank Skew-T.
    """
    skewt([1100], [0], [0], parcel = 'off')

"""Notes for development:

For bonus points, other functions could be added to calculate specific
thermodynamic quantities (i.e. potential temperature, CAPE, how high
parcel would travel until it became negatively buoyant, etc.).

Maybe also come up with way to control area of plot (i.e. focus on a
certain area of the Skew T, such as the layer between 1000 and 500mb.

Other than that, this code is in pretty good shape!"""

